#!/usr/bin/env python3
import argparse
import asyncio
import json
from pathlib import Path
from typing import Any, Dict, List

ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
import sys

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.core.agent_context import AgentContext
from app.core.agent_orchestrator import AgentOrchetrator
from app.core.metrics_tracker import metrics_tracker


OUTPUT_COLUMNS = [
    "run_id",
    "status",
    "error",
    "intent",
    "agents_planned",
    "agents_executed",
    "conversation_output",
    "symptom_matcher_output",
    "disease_info_output",
    "reasoning_output",
    "final_output",
    "relevance_score",
    "latency_ms",
    "local_db_calls",
    "local_db_successes",
    "vector_db_calls",
    "vector_db_successes",
    "internet_search_calls",
    "internet_search_successes",
    "memory_recall_calls",
    "memory_items_used",
    "memory_save_calls",
    "tool_errors",
]


def _parse_args():
    parser = argparse.ArgumentParser(
        description="Run SmartHealth multi-agent evaluation for queries from Excel and write outputs/metrics."
    )
    parser.add_argument("--input", required=True, help="Path to input Excel file (.xlsx)")
    parser.add_argument("--output", default=None, help="Output Excel path (default: <input>_evaluated.xlsx)")
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: active sheet)")
    parser.add_argument("--query-column", default="queries", help="Column header containing queries")
    parser.add_argument("--session-prefix", default="eval", help="Session ID prefix for each row run")
    parser.add_argument(
        "--create-template",
        action="store_true",
        help="Create a new Excel template with only the query column and exit.",
    )
    return parser.parse_args()


def _find_or_create_columns(ws, query_column: str) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        value = (cell.value or "").strip() if isinstance(cell.value, str) else cell.value
        if value:
            header_map[str(value)] = col_idx

    if query_column not in header_map:
        raise ValueError(f"Missing required column '{query_column}' in row 1.")

    next_col = ws.max_column + 1
    for col_name in OUTPUT_COLUMNS:
        if col_name not in header_map:
            ws.cell(row=1, column=next_col, value=col_name)
            header_map[col_name] = next_col
            next_col += 1
    return header_map


def _json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=True)


def _lookup_run_metrics(run_id: str) -> Dict[str, Any]:
    for run in metrics_tracker.get_recent_runs(limit=500):
        if run.get("run_id") == run_id:
            return run
    return {}


async def _evaluate_query(query: str, session_id: str) -> Dict[str, Any]:
    context = AgentContext(session_id=session_id)
    orchestrator = AgentOrchetrator(context)
    result = await orchestrator.run(query, capture_trace=True)
    run_id = result["run_id"]
    trace = result.get("trace", {})
    run_metrics = _lookup_run_metrics(run_id)

    agent_outputs = trace.get("agent_outputs", {})

    row_payload = {
        "run_id": run_id,
        "status": run_metrics.get("status", "ok"),
        "error": run_metrics.get("error", ""),
        "intent": trace.get("intent", run_metrics.get("intent", "")),
        "agents_planned": _json_text(trace.get("agents_planned", run_metrics.get("agents_planned", []))),
        "agents_executed": _json_text(run_metrics.get("agents_executed", [])),
        "conversation_output": _json_text(agent_outputs.get("conversation_agent")),
        "symptom_matcher_output": _json_text(agent_outputs.get("symptom_matcher_agent")),
        "disease_info_output": _json_text(agent_outputs.get("disease_info_agent")),
        "reasoning_output": _json_text(agent_outputs.get("reasoning_agent")),
        "final_output": result.get("final_output", ""),
        "relevance_score": run_metrics.get("relevance_score"),
        "latency_ms": run_metrics.get("latency_ms"),
        "local_db_calls": run_metrics.get("local_db_calls", 0),
        "local_db_successes": run_metrics.get("local_db_successes", 0),
        "vector_db_calls": run_metrics.get("vector_db_calls", 0),
        "vector_db_successes": run_metrics.get("vector_db_successes", 0),
        "internet_search_calls": run_metrics.get("internet_search_calls", 0),
        "internet_search_successes": run_metrics.get("internet_search_successes", 0),
        "memory_recall_calls": run_metrics.get("memory_recall_calls", 0),
        "memory_items_used": run_metrics.get("memory_items_used", 0),
        "memory_save_calls": run_metrics.get("memory_save_calls", 0),
        "tool_errors": run_metrics.get("tool_errors", 0),
    }
    return row_payload


async def _run(args):
    try:
        from openpyxl import Workbook, load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "Missing dependency 'openpyxl'. Install it with: pip install openpyxl"
        ) from exc

    input_path = Path(args.input).expanduser().resolve()
    if args.create_template:
        wb = Workbook()
        ws = wb.active
        if args.sheet:
            ws.title = args.sheet
        ws.cell(row=1, column=1, value=args.query_column)
        wb.save(input_path)
        print(f"Template created: {input_path}")
        return

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_evaluated.xlsx")
    )

    wb = load_workbook(input_path)
    ws = wb[args.sheet] if args.sheet else wb.active
    headers = _find_or_create_columns(ws, query_column=args.query_column)
    query_col = headers[args.query_column]

    metrics_tracker.reset()
    total_rows = ws.max_row
    processed = 0

    for row_idx in range(2, total_rows + 1):
        cell_value = ws.cell(row=row_idx, column=query_col).value
        query = str(cell_value).strip() if cell_value is not None else ""
        if not query:
            continue

        session_id = f"{args.session_prefix}-{row_idx}"
        try:
            payload = await _evaluate_query(query=query, session_id=session_id)
        except Exception as exc:
            payload = {
                "run_id": "",
                "status": "error",
                "error": str(exc),
                "intent": "",
                "agents_planned": "",
                "agents_executed": "",
                "conversation_output": "",
                "symptom_matcher_output": "",
                "disease_info_output": "",
                "reasoning_output": "",
                "final_output": "",
                "relevance_score": "",
                "latency_ms": "",
                "local_db_calls": "",
                "local_db_successes": "",
                "vector_db_calls": "",
                "vector_db_successes": "",
                "internet_search_calls": "",
                "internet_search_successes": "",
                "memory_recall_calls": "",
                "memory_items_used": "",
                "memory_save_calls": "",
                "tool_errors": "",
            }

        for col_name in OUTPUT_COLUMNS:
            ws.cell(row=row_idx, column=headers[col_name], value=payload.get(col_name))
        processed += 1

    wb.save(output_path)
    print(f"Processed rows: {processed}")
    print(f"Saved output: {output_path}")


def main():
    args = _parse_args()
    asyncio.run(_run(args))


if __name__ == "__main__":
    main()
